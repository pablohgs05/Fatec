import openpyxl

#Função para verificar se uma turma está aberta ou fechada
def turma_aberta(sheet_name):
    return "(fechada)" not in sheet_name

#Nome do arquivo Excel
arquivo_excel = "Dados Cadastrais.xlsx"

while True:
    try:
        #Carregar o arquivo
        book = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        print("O arquivo Excel não foi encontrado.")
        exit()

    #Função criada para fechar uma turma específica
    def fechar_turma(book, turma_numero):
        turma_nome = f"Turma {turma_numero}"
        if turma_nome in book.sheetnames:
            sheet = book[turma_nome]
            novo_nome = f"{turma_nome} (fechada)"
            sheet.title = novo_nome
            book.save(arquivo_excel)
            print(f"\n{turma_nome} foi fechada com sucesso.")
        else:
            print(f"\nA turma {turma_nome} já foi fechada ou não foi encontrada.")

    #Outra função que agora pergunta ao usuário se deseja continuar fechando turmas
    def continuar_fechando():
        while True:
            resposta = input("\nDeseja continuar fechando turmas? (S para continuar ou N para sair): ").strip().lower()
            if resposta == 's':
                turma_numero = int(input("Digite o número da próxima turma que deseja fechar: "))
                fechar_turma(book, turma_numero)
            elif resposta == 'n':
                book.save(arquivo_excel)
                print("Encerrando programa.")
                exit()
            else:
                print("Resposta inválida. Por favor, digite 'S' para continuar ou 'N' para sair e salvar.")

    #Função para listar turmas disponíveis
    def listar_turmas_disponiveis(book):
        abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ') and turma_aberta(sheet)]
        print("\n\n\n\n\n\nTurmas disponíveis para fechamento:", "\n")
        for i, turma in enumerate(abas_turmas, start=1):
            print(f"{turma.replace('(fechada)', '')}")
        print()

    #Pergunta ao usuário qual o número da turma que deseja fechar
    listar_turmas_disponiveis(book)
    turma_numero = int(input("Digite o número da turma que deseja fechar: "))
    fechar_turma(book, turma_numero)

    #Por fim, pergunta se gostaria de continuar fechando turmas, ativando a função
    continuar = input("\nDeseja continuar fechando turmas? (S para continuar ou N para sair): ").strip().lower()
    if continuar != 's':
        book.save(arquivo_excel)
        print("Encerrando programa.")
        break