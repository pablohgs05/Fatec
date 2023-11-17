import openpyxl
import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

#Obtenha o diretório atual em que o script Python está sendo executado
diretorio_atual = os.path.dirname(os.path.abspath(__file__ ))

#Construa o caminho completo para o arquivo Excel no diretório 'database'
caminho_arquivo_excel = os.path.join(diretorio_atual, '..', 'database', 'infodados.xlsx')

#Abrir o arquivo Excel existente ou criar um novo
if os.path.exists(caminho_arquivo_excel):
    book = openpyxl.load_workbook(caminho_arquivo_excel)
else:
    book = openpyxl.Workbook()

#Funções
def gerenciar_alunos():    
    # Função para verificar se um aluno já está alocado em uma turma
    def aluno_em_turma(planilha, aluno_chave):
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        for turma in abas_turmas:
            aba_turma = planilha[turma]
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[0].value
                cpf = row[1].value
                aluno_turma_chave = (nome, cpf)

                if aluno_turma_chave == aluno_chave:
                    return turma  # Retorna o nome da turma onde o aluno está

        return None

    # Função para listar os alunos em uma turma
    def listar_alunos_na_turma(planilha, turma_nome):
        if turma_nome in planilha.sheetnames:
            aba_turma = planilha[turma_nome]
            alunos_na_turma = []
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                nome = row[0].value
                cpf = row[1].value
                alunos_na_turma.append((nome, cpf))
            return alunos_na_turma
        else:
            return None

    # Conjunto para manter o controle dos alunos já adicionados
    alunos_adicionados = set()

    while True:
        print("\nOpções:")
        print("\n1. Adicionar alunos às turmas")
        print("2. Ver alunos disponíveis e não alocados")
        print("3. Remover aluno de uma turma")
        print("4. Voltar", "\n")

        escolha = input("Escolha uma das opções: ")

        if escolha == '1':
            #Listar as abas de turma disponíveis (apenas as abertas)
            abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

            if not abas_turmas:
                print("\nNão foram encontradas abas de turma na planilha.", "\n")
            else:
                print("\nTurmas abertas:")
                for i, turma in enumerate(abas_turmas, start=1):
                    if "(fechada)" not in turma:
                        print(f"{turma.replace('(fechada)', '')}")

                #Pergunta ao usuário em qual turma deseja adicionar os alunos
                try:
                    indice_turma = int(input("Digite o número da turma que deseja adicionar os alunos: ")) - 1

                    if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                        turma_desejada = abas_turmas[indice_turma]
                        quantidade_alunos = int(input("Quantos alunos deseja adicionar: "))

                        #Verificar alunos na aba "Alunos" que não foram colocados em nenhuma turma
                        aba_alunos = book['Cadastro']
                        alunos_disponiveis = []

                        for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
                            nome = row[0].value
                            cpf = row[1].value
                            funcao = row[3].value
                            aluno_chave = (nome, cpf)

                            if not aluno_em_turma(book, aluno_chave) and funcao == "aluno":
                                alunos_disponiveis.append((nome, cpf, row[2].value))

                        #Selecionar a quantidade desejada de alunos disponíveis
                        alunos_selecionados = alunos_disponiveis[:quantidade_alunos]

                        # Adicionar os alunos selecionados à turma
                        aba_turma = book[turma_desejada]
                        for aluno in alunos_selecionados:
                            nome, cpf, email = aluno
                            nova_linha = [nome, cpf, email]
                            linha = [nome, cpf, email]
                            aba_turma.append(linha)
                            print(f"Aluno {nome} adicionado à {turma_desejada} com sucesso.")
                            alunos_adicionados.add(aluno)

                    else:
                        print("\nTurma não encontrada ou está fechada.")
                except ValueError:
                    print("\nDigite um número válido para selecionar a turma.")

        elif escolha == '2':
            alunos_nao_alocados = 0

            #Verificar alunos na aba "alunos" que não foram alocados a nenhuma turma
            aba_alunos = book['Cadastro']
            for row in aba_alunos.iter_rows(min_row=2, max_row=aba_alunos.max_row, min_col=1, max_col=4):
                nome = row[0].value
                cpf = row[1].value
                funcao = row[3].value
                aluno_chave = (nome, cpf)

                if not aluno_em_turma(book, aluno_chave) and funcao == "aluno":
                    alunos_nao_alocados += 1

            print(f"\nAlunos disponíveis para alocação: {alunos_nao_alocados}")

        elif escolha == '3':
            # Remover aluno de uma turma
            turma_escolhida = input("Digite o nome da turma da qual deseja remover o aluno: ")
            alunos_na_turma = listar_alunos_na_turma(book, turma_escolhida)

            if alunos_na_turma:
                print(f"Alunos na turma {turma_escolhida}:")
                for i, aluno in enumerate(alunos_na_turma, start=1):
                    nome, cpf = aluno
                    print(f"{i}. {nome} (CPF: {cpf})")

                escolha_aluno = input("Digite o número do aluno que deseja remover: ")

                try:
                    escolha_aluno = int(escolha_aluno)
                    if 1 <= escolha_aluno <= len(alunos_na_turma):
                        aluno_remover = alunos_na_turma[escolha_aluno - 1]
                        nome_aluno, cpf_aluno = aluno_remover

                        turma_do_aluno = aluno_em_turma(book, (nome_aluno, cpf_aluno))
                        if turma_do_aluno:
                            aba_turma = book[turma_do_aluno]
                            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=1, max_col=3):
                                nome = row[0].value
                                cpf = row[1].value
                                if (nome, cpf) == (nome_aluno, cpf_aluno):
                                    aba_turma.delete_rows(row[0].row)
                                    print(f"Aluno {nome_aluno} removido da turma {turma_do_aluno}.")
                                    break
                        else:
                            print("Aluno não encontrado em nenhuma turma.")
                    else:
                        print("Escolha de aluno inválida.")
                except ValueError:
                    print("Escolha de aluno inválida.")
            else:
                print("Turma não encontrada ou não há alunos nessa turma.")

        elif escolha == '4':
            break
        else:
            print("\nOpção inválida. Escolha 1 para adicionar alunos, 2 para verificar alunos disponíveis, 3 para remover um aluno de uma turma, ou 4 para sair.")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)
def gerenciar_professores():
    # Função para verificar se um professor já está alocado em uma turma
    def professor_em_turma(planilha, professor_chave):
        abas_turmas = [sheet for sheet in planilha.sheetnames if sheet.startswith('Turma ')]

        for turma in abas_turmas:
            aba_turma = planilha[turma]
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                nome = row[0].value
                cpf = row[1].value
                professor_turma_chave = (nome, cpf)

                if professor_turma_chave == professor_chave:
                    return turma  # Retorna o nome da turma onde o professor está

        return None

    # Função para listar os professores em uma turma
    def listar_professores_na_turma(planilha, turma_nome):
        if turma_nome in planilha.sheetnames:
            aba_turma = planilha[turma_nome]
            professores_na_turma = []
            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                nome = row[0].value
                cpf = row[1].value
                professores_na_turma.append((nome, cpf))
            return professores_na_turma
        else:
            return None

    # Conjunto para manter o controle dos professores já adicionados
    professores_adicionados = set()

    while True:
        print("\nOpções:")
        print("\n1. Adicionar professores às turmas")
        print("2. Ver professores disponíveis e não alocados")
        print("3. Remover professor de uma turma")
        print("4. Voltar", "\n")

        escolha = input("Escolha uma das opções: ")

        if escolha == '1':
            # Listar as abas de turma disponíveis (apenas as abertas)
            abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ')]

            if not abas_turmas:
                print("\nNão foram encontradas abas de turma na planilha.", "\n")
            else:
                print("\nTurmas abertas:")
                for i, turma in enumerate(abas_turmas, start=1):
                    if "(fechada)" not in turma:
                        print(f"{turma.replace('(fechada)', '')}")

                # Pergunta ao usuário em qual turma deseja adicionar os professores
                try:
                    indice_turma = int(input("Digite o número da turma que deseja adicionar os professores: ")) - 1

                    if 0 <= indice_turma < len(abas_turmas) and "(fechada)" not in abas_turmas[indice_turma]:
                        turma_desejada = abas_turmas[indice_turma]
                        quantidade_professores = int(input("Quantos professores deseja adicionar: "))

                        # Verificar professores na aba "Cadastro" que não foram colocados em nenhuma turma
                        aba_professores = book['Cadastro']
                        professores_disponiveis = []

                        for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
                            nome = row[0].value
                            cpf = row[1].value
                            funcao = row[3].value
                            professor_chave = (nome, cpf)

                            if not professor_em_turma(book, professor_chave) and funcao == "professor":
                                professores_disponiveis.append((nome, cpf))

                        # Selecionar a quantidade desejada de professores disponíveis
                        professores_selecionados = professores_disponiveis[:quantidade_professores]

                        # Função para adicionar um professor a uma turma específica
                        def adicionar_professor_a_turma(planilha, turma_destino, professor_nome, professor_cpf):
                            aba_turma = planilha[turma_destino]

                            # Encontrar a primeira linha vazia após os cabeçalhos "Professores" e "CPF - Prof"
                            primeira_linha_vazia = 2  # Começando na linha 2 para evitar os cabeçalhos

                            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                                if row[0].value is None or row[1].value is None:
                                    break
                                primeira_linha_vazia += 1


                            # Agora, você pode adicionar os dados do professor na primeira linha vazia
                            aba_turma.cell(row=primeira_linha_vazia, column=4).value = professor_nome
                            aba_turma.cell(row=primeira_linha_vazia, column=5).value = professor_cpf

                            # Salvar o livro após adicionar o professor
                            planilha.save(caminho_arquivo_excel)

                        # Adicionar os professores selecionados à turma
                        aba_turma = book[turma_desejada]
                        for professor in professores_selecionados:
                            nome, cpf = professor
                            adicionar_professor_a_turma(book, turma_desejada, nome, cpf)
                            print(f"Professor {nome} adicionado à {turma_desejada} com sucesso.")
                            professores_adicionados.add(professor)

                        else:
                            print("\nTurma não encontrada ou está fechada.")
                except ValueError:
                    print("\nDigite um número válido para selecionar a turma.")

        elif escolha == '2':
            professores_nao_alocados = 0

            # Verificar professores na aba "Cadastro" que não foram alocados a nenhuma turma
            aba_professores = book['Cadastro']
            for row in aba_professores.iter_rows(min_row=2, max_row=aba_professores.max_row, min_col=1, max_col=4):
                nome = row[0].value
                cpf = row[1].value
                funcao = row[3].value
                professor_chave = (nome, cpf)

                if not professor_em_turma(book, professor_chave) and funcao == "professor":
                    professores_nao_alocados += 1

            print(f"\nProfessores disponíveis para alocação: {professores_nao_alocados}")

        elif escolha == '3':
            # Remover professor de uma turma
            turma_escolhida = input("Digite o nome da turma da qual deseja remover o professor: ")
            professores_na_turma = listar_professores_na_turma(book, turma_escolhida)

            if professores_na_turma:
                print(f"Professores na turma {turma_escolhida}:")
                for i, professor in enumerate(professores_na_turma, start=1):
                    nome, cpf = professor
                    print(f"{i}. {nome} (CPF: {cpf})")

                escolha_professor = input("Digite o número do professor que deseja remover: ")

                try:
                    escolha_professor = int(escolha_professor)
                    if 1 <= escolha_professor <= len(professores_na_turma):
                        professor_remover = professores_na_turma[escolha_professor - 1]
                        nome_professor, cpf_professor = professor_remover

                        turma_do_professor = professor_em_turma(book, (nome_professor, cpf_professor))
                        if turma_do_professor:
                            aba_turma = book[turma_do_professor]
                            for row in aba_turma.iter_rows(min_row=2, max_row=aba_turma.max_row, min_col=4, max_col=5):
                                nome = row[0].value
                                cpf = row[1].value
                                if (nome, cpf) == (nome_professor, cpf_professor):
                                    aba_turma.delete_rows(row[0].row)
                                    print(f"Professor {nome_professor} removido da turma {turma_do_professor}.")
                                    break
                        else:
                            print("Professor não encontrado em nenhuma turma.")
                except ValueError:
                    print("Escolha de professor inválida.")

        elif escolha == '4':
            break
        else:
            print("\nOpção inválida. Escolha 1 para adicionar professores, 2 para verificar professores disponíveis, 3 para remover um professor de uma turma, ou 4 para sair.")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)
def grupos():
    # Função para criar grupos em uma turma
    def criar_grupos(planilha, turma_nome, num_alunos_por_grupo):
        # Carregar o arquivo da planilha
        wb = openpyxl.load_workbook(planilha)
        
        # Selecionar a aba da turma
        try:
            sheet = wb[turma_nome]
        except KeyError:
            print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
            return
        
        # Obter o índice da coluna que contém o cabeçalho "Grupos"
        coluna_grupos = None
        for cell in sheet[1]:  # Percorre as células da primeira linha
            if cell.value == "Grupos":
                coluna_grupos = cell.column_letter
                break
        
        if coluna_grupos is None:
            print("Não foi encontrada uma coluna com o cabeçalho 'Grupos'.")
            return
        
        # Obter a lista de alunos na turma
        alunos = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row if cell.value]
        
        # Exibir a quantidade de alunos na turma
        num_alunos = len(alunos)
        print(f"\nA turma '{turma_nome}' possui {num_alunos} alunos.")
        
        # Calcular o número total de grupos
        num_grupos = num_alunos // num_alunos_por_grupo
        alunos_restantes = num_alunos % num_alunos_por_grupo
        
        # Criar grupos com o número especificado de alunos
        grupos = []
        for i in range(num_grupos):
            grupo_alunos = alunos[i * num_alunos_por_grupo: (i + 1) * num_alunos_por_grupo]
            grupos.append(f'Grupo {i + 1} - {", ".join(grupo_alunos)}')
        
        # Se houver alunos restantes, criar o último grupo
        if alunos_restantes > 0:
            grupo_alunos = alunos[-alunos_restantes:]
            grupos.append(f'Grupo {num_grupos + 1} - {", ".join(grupo_alunos)}')
        
        # Adicionar os grupos à coluna "Grupos" na planilha
        for i, grupo in enumerate(grupos):
            sheet[f"{coluna_grupos}{i + 2}"] = grupo
        
        # Salvar a planilha
        wb.save(planilha)
        print(f"\nForam criados {len(grupos)} grupos com {num_alunos_por_grupo} alunos cada{' e um grupo final com ' + str(alunos_restantes) + ' alunos' if alunos_restantes > 0 else ''} na '{turma_nome}'.")

    # Função para listar as turmas existentes
    def listar_turmas(planilha):
        wb = openpyxl.load_workbook(planilha)
        turmas = [sheet for sheet in wb.sheetnames if sheet.startswith('Turma ')]
        if turmas:
            print("\nTurmas existentes:", "\n")
            for turma in turmas:
                print(turma)
        else:
            print("\nNão foram encontradas abas de turma na planilha.")

    # Função para contar alunos em uma turma
    def contar_alunos(planilha, turma_nome):
        # Carregar o arquivo da planilha
        wb = openpyxl.load_workbook(planilha)
        
        # Selecionar a aba da turma
        try:
            sheet = wb[turma_nome]
        except KeyError:
            print(f"\nA '{turma_nome}' não foi encontrada na planilha.")
            return
        
        # Contar alunos na turma
        num_alunos = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value)
        return num_alunos  # Retornar a quantidade de alunos na turma

    # Função principal do programa
    def main():
        while True:
            print("\nOpções:")
            print("\n1. Criar grupos")
            print("2. Listar turmas existentes")
            print("3. Voltar", "\n")
            
            escolha = input("Escolha uma das opções: ")
            
            if escolha == '1':
                turma = input('\nDigite o nome da turma: ')
                alunos_na_turma = contar_alunos(caminho_arquivo_excel, turma)  # Mostrar a quantidade de alunos na turma
                if alunos_na_turma is not None:
                    if alunos_na_turma > 0:
                        print(f"\nA turma '{turma}' possui {alunos_na_turma} alunos.")
                        alunos_por_grupo = int(input("\nDigite o número de alunos por grupo: "))
                        criar_grupos(caminho_arquivo_excel, turma, alunos_por_grupo)
            elif escolha == '2':
                listar_turmas(caminho_arquivo_excel)
            elif escolha == '3':
                break
            else:
                print("\nOpção inválida. Tente novamente.", "\n")

    if __name__ == "__main__":
        main()
def turmas():
    # Função para criar uma nova aba de turma
    def criar_nova_turma(book, numero_turma):
        nova_aba = book.create_sheet(f"Turma {numero_turma}")
        cabecalhos = ["Alunos", "CPF", "Email", "Professores", "CPF - Prof", "Grupos", "Inicio do Curso", "Fim do Curso"]
        nova_aba.append(cabecalhos)

        # Ajustar a largura da coluna e centralizar os cabeçalhos
        for col_idx, header in enumerate(cabecalhos, 1):
            coluna_letra = get_column_letter(col_idx)
            cell = nova_aba[f"{coluna_letra}1"]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
            cell.value = header
            nova_aba.column_dimensions[coluna_letra].width = 30

    # Função para exibir o número de turmas
    def mostrar_numero_de_turmas(book):
        turmas_existentes = [sheet for sheet in book.sheetnames if sheet.startswith("Turma ")]
        if not turmas_existentes:
            print("\nNenhuma turma encontrada.")
        else:
            print("\nTurmas existentes:")
            for turma in turmas_existentes:
                print("\n", turma)
            print(f"\nTotal de turmas: {len(turmas_existentes)}")

    # Função para excluir uma ou mais turmas
    def excluir_turmas(book, turmas_a_excluir):
        for turma_nome in turmas_a_excluir:
            if turma_nome in book.sheetnames:
                book.remove(book[turma_nome])
                print(f"\nTurma {turma_nome} excluída com sucesso.")
            else:
                print(f"\nA turma {turma_nome} não foi encontrada.")
        book.save(caminho_arquivo_excel)

    # Função para fechar turmas
    def fechar_turmas():
        while True:
            try:
                # Carrega o arquivo
                book = openpyxl.load_workbook(caminho_arquivo_excel)
            except FileNotFoundError:
                print("O banco de dados não foi encontrado.")
                exit()

            def fechar_turma(book, turma_numero):
                turma_nome = f"Turma {turma_numero}"
                if turma_nome in book.sheetnames:
                    sheet = book[turma_nome]
                    novo_nome = f"{turma_nome} (fechada)"
                    sheet.title = novo_nome
                    book.save(caminho_arquivo_excel)  # Salva as alterações no arquivo Excel
                    print(f"\n{turma_nome} foi fechada com sucesso.")
                    return True
                else:
                    print(f"\n\n\n\n\n\n\n\nDesculpe, a {turma_nome} já foi fechada ou não foi encontrada, tente novamente.")
                    return False

            def listar_turmas_disponiveis(book):
                abas_turmas = [sheet for sheet in book.sheetnames if sheet.startswith('Turma ') and not sheet.endswith('(fechada)')]
                print("\nTurmas disponíveis para fechamento:")
                for i, turma in enumerate(abas_turmas, start=1):
                    print(f"{turma.replace('(fechada)', '')}")
                print()

            listar_turmas_disponiveis(book)
            turma_numero = int(input("Digite o número da turma que deseja fechar: "))

            if fechar_turma(book, turma_numero):
                while True:
                    continuar = input("\nDeseja continuar fechando turmas? (S para continuar ou N para sair): ").strip().lower()
                    if continuar == 's':
                        break
                    elif continuar == 'n':
                        book.save(caminho_arquivo_excel)  # Salva as alterações antes de sair
                        exit()
                    else:
                        print("\n\nResposta inválida, por favor, tente novamente.")

    # Loop principal
    while True:
        print("\nOpções:")
        print("\n1. Criar nova turma")
        print("2. Visualizar turmas")
        print("3. Excluir turmas")
        print("4. Fechar turmas")
        print("5. Voltar", "\n")
        
        opcao = input('Escolha uma das opções: ')
        
        if opcao == "1":
            quantidade_turmas = int(input("Digite a quantidade de turmas que deseja criar: "))
            for _ in range(quantidade_turmas):
                proxima_turma = sum(1 for sheet in book.sheetnames if sheet.startswith("Turma ")) + 1
                criar_nova_turma(book, proxima_turma)
                print(f"\nTurma {proxima_turma} criada com sucesso.")
        elif opcao == "2":
            mostrar_numero_de_turmas(book)
        elif opcao == "3":
            turmas_a_excluir = input("Digite o nome das turmas que deseja excluir: ")
            turmas_a_excluir = [turma.strip() for turma in turmas_a_excluir.split(",")]
            excluir_turmas(book, turmas_a_excluir)
        elif opcao == "4":
            fechar_turmas()

        elif opcao == "5":
            break
        else:
            print("\nOpção inválida.", "\n")


    # Salve as alterações no arquivo Excel
    book.save(caminho_arquivo_excel)

#Menu
while True:
    try:
        print("\n\n\nOpções:")
        print("\n1. Associação de alunos e professores")
        print("2. Grupos")
        print("3. Criação de turmas")
        print("4. Sair do programa", "\n")
        escolha = int(input("Escolha uma das opções: "))

        while True:
            if escolha == 1:
                print("\n\nOpções:")
                print("\n1. Gerenciar alunos")
                print("2. Gerenciar professores")
                print("3. Voltar", "\n")
                escolha1 = input("Escolha uma das opções: ")
                if escolha1 == '1':
                    gerenciar_alunos()
                elif escolha1 == '2':
                    gerenciar_professores()
                elif escolha1 == '3':
                    break
                else:
                    print("\nOpção inválida, tente novamente.")
            elif escolha == 2:
                print("\n\nOpções:")
                print("\n1. Gerenciar grupos")
                print("2. Voltar", "\n")
                escolha2 = input("Escolha uma das opções: ")
                if escolha2 == '1':
                    grupos()
                elif escolha2 == '2':
                    break
                else:
                    print("\nOpção inválida, tente novamente.")
            elif escolha == 3:
                print("\n\nOpções:")
                print("\n1. Gerenciar turmas")
                print("2. Voltar", "\n")
                escolha3 = input("Escolha uma das opções: ")
                if escolha3 == '1':
                    turmas()
                elif escolha3 == '2':
                    break
                else:
                    print("\nOpção inválida, tente novamente.")
            elif escolha == 4:
                print("\n\nAlterações realizadas com sucesso, encerrando o programa.", "\n")
                exit()
            else:
                print("\n\n\n\nOpção inválida, tente novamente.")
                break
    except ValueError:
        print("\n\nPor favor, entre apenas com números.")

#Salva as alterações no arquivo Excel
book.save(caminho_arquivo_excel)